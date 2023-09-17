import httpx
import json
import asyncio
import time
import aiohttp
import xlsxwriter
import pickle
import openpyxl
from pathlib import Path

JSON_LD_HEADER = {'Accept': 'Application/ld+json'}

async def gather_with_concurrency(n, *tasks):
    semaphore = asyncio.Semaphore(n)

    async def sem_task(task):
        async with semaphore:
            return await task

    return await asyncio.gather(*(sem_task(task) for task in tasks))

async def get_nav_list_data_async(url):

    conn = aiohttp.TCPConnector(limit=None, ttl_dns_cache=300)

    async with aiohttp.ClientSession(connector=conn) as session:
        async with session.get(url["item"]["@id"], headers=JSON_LD_HEADER) as response:
            return await response.read()


async def get_entries_from_nav_list(nav_list):
    s = time.perf_counter()
    resps = await gather_with_concurrency(90, *map(get_nav_list_data_async,
                                                   nav_list[
                                                       "itemListElement"]))
    data = [json.loads(resp) for resp in resps]
    elapsed = time.perf_counter() - s
    print(f"{elapsed}s")
    return data


def read_from_excel(file_path: Path, sheet_name, column_name):
    wb = openpyxl.load_workbook(filename=file_path, data_only=True)
    ws = wb[sheet_name]
    return [c.value for c in ws[column_name]][1:]


def get_ecoinvent_entries_by_name(entry_name: str):
    http_top_level_nav = httpx.get('https://glossary.ecoinvent.org', headers=JSON_LD_HEADER)
    top_level_nav = json.loads(http_top_level_nav.text)

    entry_url = None
    for item in top_level_nav["itemListElement"]:
        if "ld-schema" in item["item"]["name"]:
            continue

        if entry_name in item["item"]["name"]:
            entry_url = item["item"]["url"]
            break

    if entry_url is None:
        raise ValueError(f"{entry_name} is not a valid entry name")

    http_intermediate_exchange_nav = httpx.get(entry_url, headers=JSON_LD_HEADER)
    intermediate_exchange_nav = json.loads(http_intermediate_exchange_nav.text)
    loop = asyncio.get_event_loop()
    all_int_ex = loop.run_until_complete(get_entries_from_nav_list(intermediate_exchange_nav))

    return all_int_ex

def write_list_to_excel(data, file_path: Path, sheet_name):

    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet(sheet_name)

    for i, key in enumerate(data[0].keys()):
        worksheet.write(0, i, key)

    for row_num, row_data in enumerate(data, start=1):
        for col_num, (key, value) in enumerate(row_data.items()):
            worksheet.write(row_num, col_num, value)

    workbook.close()

def get_dictionary_from_list(entry_list, key):
    return {entry[key]: entry for entry in entry_list}


if __name__ == "__main__":
    soup = read_from_excel(Path("../data/transport.xlsx"), "transport", "I")
    _id = read_from_excel(Path("../data/transport.xlsx"), "transport", "A")
    ids_soup = [x for x in zip(_id, soup) if x[1] is not None]

    units = get_ecoinvent_entries_by_name("unit")
    write_list_to_excel(units, Path("../data/units.xlsx"), "units")

    units_dict = get_dictionary_from_list(units, "slug")
    print(units_dict)

    # int_ex = asyncio.get_event_loop().run_until_complete(hestia_api())


